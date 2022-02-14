#!/usr/bin/python
# -*- coding: UTF-8 -*-

"""
Author Xing GAO
https://github.com/NatGaoXing
这是一个简约的抽奖系统，拥有最基础的共嗯那个。并不美观 :)
开发时间实在过短，没有什么注释，能用就行HHH
"""
from openpyxl import load_workbook
import random
from tkinter import *
from PIL import Image, ImageTk
import time


def _from_rgb(rgb):
    return "#%02x%02x%02x" % rgb


class A:

    def __init__(self, master=None):
        self.root = master
        self.root.geometry('800x600+200+100')
        self.root.title('寻味抽奖活动页面 Version 1.0.2')
        self.frm1 = Frame(self.root)
        self.frm2 = Frame(self.root)
        self.frm3 = Frame(self.root)
        self.num = 0
        self.createpage()

    def button_1_call_back(self):
        self.frm1.destroy()
        self.frm1 = Frame(self.root)
        self.frm2 = Frame(self.root)
        self.frm3 = Frame(self.root)
        self.createpage()

        Label(self.frm1, text="获得寻味倾心特别奖的是: ", height=1, width=20).place(x=220, y=225)

        for k in range(2, max_row + 1):
            Label(self.frm1, text=ws['B' + str(k)].value, height=1, width=20).place(x=220, y=250)
            root.update()
            self.num = self.num + 1
            time.sleep(2 / max_row)

        Label(self.frm1, text=str(name_list_price[-1]), height=1, width=20).place(x=220, y=250)

    def button_2_call_back(self):
        self.frm1.destroy()
        self.frm1 = Frame(self.root)
        self.frm2 = Frame(self.root)
        self.frm3 = Frame(self.root)
        self.createpage()

        Label(self.frm1, text="获得八折券的是: ", height=1, width=20).place(x=220, y=175)
        for k in range(2, max_row + 1):
            for i in range(2):
                for j in range(4):
                    if k > max_row:
                        Label(self.frm1, text=ws['B' + str(k + i * 5 + j)].value, height=1, width=20).place(
                            x=7 + 147 * j, y=50 + 100 * i)
                    else:
                        Label(self.frm1, text=ws['B' + str(random.randint(2, max_row))].value, height=1,
                              width=20).place(x=7 + 147 * j, y=200 + 100 * i)

            root.update()
            self.num = self.num + 1
            time.sleep(2 / max_row)

        for i in range(2):
            for j in range(4):
                Label(self.frm1, text=str(name_list_price[i * 4 + j + 14]), height=1, width=20).place(x=7 + 147 * j, y=200 + 100 * i)

        Label(self.frm1, text="获得免费名额的是: ", height=1, width=20).place(x=220, y=25)
        for k in range(2, max_row + 1):
            Label(self.frm1, text=ws['B' + str(k)].value, height=1, width=20).place(x=220, y=50)
            root.update()
            self.num = self.num + 1
            time.sleep(2 / max_row)

        Label(self.frm1, text=str(name_list_price[-2]), height=1, width=20).place(x=220, y=50)

    def button_3_call_back(self):

        self.frm1.destroy()
        self.frm1 = Frame(self.root)
        self.frm2 = Frame(self.root)
        self.frm3 = Frame(self.root)
        self.createpage()

        # Label(self.frm1, text="获得三等特奖的是: ", height=1, width=20).place(x=20, y=10)
        for k in range(2, max_row + 1):
            for i in range(5):
                for j in range(3):
                    if k > max_row:
                        Label(self.frm1, text=ws['B' + str(k + i * 5 + j)].value, height=1, width=20).place(x=20 + 210 * j, y=50 + 100 * i)
                    else:
                        Label(self.frm1, text=ws['B' + str(random.randint(2, max_row))].value, height=1, width=20).place(x=20 + 210 * j, y=50 + 100 * i)

            root.update()
            self.num = self.num + 1
            time.sleep(2 / max_row)

        for i in range(5):
            for j in range(3):
                Label(self.frm1, text=str(name_list_price[i * 4 + j + 0]), height=1, width=20).place(x=20 + 210 * j, y=50 + 100 * i)

    def createpage(self):

        self.frm1.config(bg='blue', height=500, width=600)
        self.frm1.place(x=180, y=90)

        self.frm2.config(bg='red', height=500, width=155)
        self.frm2.place(x=20, y=90)

        self.frm3.config(bg='#FF97A7', height=80, width=760)
        self.frm3.place(x=20, y=5)

        # frm3下的Label
        Label(self.frm3, bg='#FF97A7', text='寻味情人节给你特别的宠爱',
              fg='black', font='Arial 18 bold').place(x=230, y=25)
        # frm2下的Button
        load = Image.open('bg2.png')
        render = ImageTk.PhotoImage(load)
        img = Label(self.frm2, image=render)
        img.image = render
        img.place(x=0, y=0, height=500, width=155)

        load = Image.open('XW.png')
        render = ImageTk.PhotoImage(load)
        img = Label(self.frm2, image=render)
        img.image = render
        img.place(x=37, y=335, height=80, width=80)
        Button(self.frm2, text='寻味倾心特别奖\n兰蔻情人节日历', command=self.button_1_call_back).place(x=28, y=420, width=100)

        load = Image.open('YYL.png')
        render = ImageTk.PhotoImage(load)
        img = Label(self.frm2, image=render)
        img.image = render
        img.place(x=37, y=175, height=80, width=80)
        Button(self.frm2, text='幺幺灵剧本杀\n免费游戏&八折券', command=self.button_2_call_back).place(x=27, y=260, width=100)

        load = Image.open('XFT.png')
        render = ImageTk.PhotoImage(load)
        img = Label(self.frm2, image=render)
        img.image = render
        img.place(x=37, y=15, height=80, width=80)
        Button(self.frm2, text='幸福堂\n免费奶茶兑换券', command=self.button_3_call_back).place(x=27, y=100, width=100)

        # frm1下的控件
        load = Image.open('bg1.png')
        render = ImageTk.PhotoImage(load)
        img = Label(self.frm1, image=render)
        img.image = render
        img.place(x=0, y=0, height=500, width=600)


if __name__ == '__main__':

    wb = load_workbook('情人节参与抽奖活动名单.xlsx')
    ws = wb['Sheet1']

    wb2 = load_workbook('获奖名单.xlsx')
    ws2 = wb2['Sheet1']

    max_row = 101 # 具体视名单而定

    if max_row > 25:
        for j in range(1, max_row + 1):
            for k in range(1, 3 + 1):
                if k == 0:
                    print(str(ws['A' + str(j)].value) + '\t', end='')
                elif k == 1:
                    print(str(ws['B' + str(j)].value) + '\t', end='')
                else:
                    print(str(ws['C' + str(j)].value) + '\t', end='')
            print()

        name_list_price = []
        while len(name_list_price) < 15:
            random_number = random.randint(2, max_row)
            if len(name_list_price) > 0:
                # 检验重复
                i = 1
                is_price = False
                while i < len(name_list_price):
                    if name_list_price[i] == str(ws['B' + str(random_number)].value):
                        is_price = True
                        i = len(name_list_price)
                    i = i + 1
                if not is_price:
                    name_list_price.append(str(ws['B' + str(random_number)].value))
            else:
                name_list_price.append(str(ws['B' + str(random_number)].value))

        while len(name_list_price) < 24:
            random_number = random.randint(2, max_row)
            if len(name_list_price) > 0:
                # 检验重复
                i = 1
                is_price = False
                while i < len(name_list_price):
                    if name_list_price[i] == str(ws['B' + str(random_number)].value):
                        is_price = True
                        i = len(name_list_price)
                    i = i + 1
                if not is_price:
                    name_list_price.append(str(ws['B' + str(random_number)].value))
            else:
                name_list_price.append(str(ws['B' + str(random_number)].value))

        while len(name_list_price) < 25:
            random_number = random.randint(2, max_row)
            if len(name_list_price) > 0:
                # 检验重复
                i = 1
                is_price = False
                while i < len(name_list_price):
                    if name_list_price[i] == str(ws['B' + str(random_number)].value):
                        is_price = True
                        i = len(name_list_price)
                    i = i + 1
                if not is_price:
                    name_list_price.append(str(ws['B' + str(random_number)].value))
            else:
                name_list_price.append(str(ws['B' + str(random_number)].value))

        list_15 = []
        for i in range(15):
            list_15.append(name_list_price[i])
        list_9_8 = []
        for i in range(15, 23):
            list_9_8.append(name_list_price[i])

        print("15人奖项获奖名单：")
        print(list_15)
        print("9-8人奖项获奖名单：")
        print(list_9_8)
        print("9-1人奖项获奖名单：")
        print(name_list_price[-2])
        print("1人奖项获奖名单：")
        print(name_list_price[-1])

        # print(random_number)
        # print(str(ws['B' + str(random_number)].value) + ' ' + str(ws['C' + str(random_number)].value))
    else:
        print("人数不足25人。")

    root = Tk()
    A(root)
    mainloop()

